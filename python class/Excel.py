import pandas , os , random
from faker import Faker
from openpyxl import *
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart3D
path = r'Data\EmployeeManagement.xlsx'
if not os.path.exists(path):
    workBook = Workbook()
    workBook[workBook.sheetnames[0]].title = 'Data1'
    workBook.save(path)
workBook = load_workbook(path)
# workBook.create_sheet('New Next Sheet')
workSheet = workBook.active
if 'Sales_Table' not in workSheet.tables.keys():
    workSheet.append(['Sales ID','Company Email','Quantity','Price','Item','Brand','Region'])
    # genderValidater = DataValidation(type='list',formula1='"Male,Female"',allow_blank=True)
    # workSheet.add_data_validation(genderValidater)
    # genderValidater.add('E2:E1048576')
    table = Table(displayName='Sales_Table',ref='A1:G1')
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium8',showFirstColumn=False,showLastColumn=False,showColumnStripes=True,showRowStripes=True)
    workSheet.add_table(table)
else:table = workSheet.tables['Sales_Table']
# workBook.create_sheet('New Next Sheet')
# workSheet.title = 'First sheet'
# workSheet = workBook['New Next Sheet']
# workBook.active = workBook.index(workSheet)
items = ['Laptop','Mouse','Keyboard','TV','Mobile','GPU','Headset','Headphones','Mic','USB Port']
region = ['North','South','East','West']
brand = ['Samsung','HP','Zebronics','Vivo','Acer','Panasonic','Sony']
generator = Faker()
for i in range(75):workSheet.append([generator.iban(),generator.company_email(),generator.random_int(1,20),generator.random_number(digits=5),random.choice(items),random.choice(brand),random.choice(region)])
# for i in workSheet.iter_rows():print(i[0].value)
table.ref = f'A1:{get_column_letter(workSheet.max_column)}{workSheet.max_row}'
# workSheet.delete_rows(1,4)
# workSheet['F2'].comment = Comment(text='This is a Comment',author='Nikish Daniel')
workBook.save(path)
workBook.close()