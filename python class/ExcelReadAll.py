import openpyxl
from collections import defaultdict

sheetsData = defaultdict(list)
def fetch():
    path = 'Data/EmployeeManagement.xlsx'
    import os
    if not os.path.exists(path):return 'File not Found.'
    workBook = openpyxl.load_workbook(path)
    for i in workBook.sheetnames:
        workSheet = workBook[i]
        workBook.active = workSheet
        currentSheetTable = list(workSheet.tables.keys())[0]
        for i in workSheet.iter_rows():sheetsData[currentSheetTable].append([x.value for x in i])
    print(sheetsData)
fetch()